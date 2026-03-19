#!/usr/bin/env python
"""Convert one or more benchmark JSON files into a master Excel workbook.

Each JSON file becomes a source label. The workbook contains:
    raw         — every individual run record, one row each
    summary     — mean / stdev / min / max per (label, case_id)
    metadata    — environment info per label (hostname, os, python, etc.)

Usage:
    uv run --with openpyxl --with pandas scripts/benchmark_to_excel.py \
        results_current.json results_compare.json \
        --out master.xlsx

    # single file
    uv run --with openpyxl --with pandas scripts/benchmark_to_excel.py \
        results_current.json --out baseline.xlsx
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path

# ---------------------------------------------------------------------------
# Column definitions — order and display names for every sheet
# ---------------------------------------------------------------------------

RAW_COLUMNS = [
    # identity
    ("label",           "label"),
    ("case_id",         "case_id"),
    ("fn",              "function"),
    ("packages",        "packages"),
    ("month",           "month"),
    ("start_month",     "start_month"),
    ("end_month",       "end_month"),
    ("monthly",         "monthly"),
    ("run_index",       "run_index"),
    ("started_at",      "started_at"),
    # timing
    ("wall_ms",         "wall_ms"),
    ("import_ms",       "import_ms"),
    ("call_ms",         "call_ms"),
    # network
    ("bytes_recv",      "bytes_recv"),
    ("bytes_sent",      "bytes_sent"),
    ("packets_recv",    "packets_recv"),
    ("packets_sent",    "packets_sent"),
    # derived network (computed here, not in the runner)
    ("kb_recv",         "kb_recv"),
    ("mb_recv",         "mb_recv"),
    # memory
    ("rss_before_mb",   "rss_before_mb"),
    ("rss_after_mb",    "rss_after_mb"),
    ("rss_delta_mb",    "rss_delta_mb"),
    # cpu
    ("cpu_user_ms",     "cpu_user_ms"),
    ("cpu_system_ms",   "cpu_system_ms"),
    # result
    ("rows",            "rows"),
    # versions
    ("pkg_version",     "pkg_version"),
    ("python_version",  "python_version"),
    # status
    ("error",           "error"),
]

SUMMARY_METRICS = [
    "wall_ms", "import_ms", "call_ms",
    "bytes_recv", "kb_recv", "mb_recv",
    "rss_delta_mb", "cpu_user_ms", "cpu_system_ms",
    "rows",
]

METADATA_COLUMNS = [
    "label", "schema_version", "started_at", "finished_at",
    "hostname", "os", "cpu_count", "python_version",
    "source_file", "total_records", "error_records",
]

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------

HEADER_BG   = "2C2C2A"
HEADER_FG   = "FFFFFF"
ALT_ROW_BG  = "F1EFE8"
ERROR_BG    = "FCEBEB"
ERROR_FG    = "A32D2D"
GREEN_BG    = "EAF3DE"
AMBER_BG    = "FAEEDA"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def load_json(path: Path) -> dict:
    return json.loads(path.read_text())


def flatten_record(rec: dict) -> dict:
    """Add derived columns and normalise list fields."""
    out = dict(rec)
    out["packages"]  = ", ".join(rec.get("packages") or [])
    out["kb_recv"]   = round(rec.get("bytes_recv", 0) / 1024,         3)
    out["mb_recv"]   = round(rec.get("bytes_recv", 0) / 1024 / 1024,  3)
    return out


def _stats(values: list[float]) -> dict:
    if not values:
        return dict(mean=None, stdev=None, min=None, max=None, count=0)
    n    = len(values)
    mean = sum(values) / n
    var  = sum((v - mean) ** 2 for v in values) / n if n > 1 else 0.0
    return dict(
        mean  = round(mean,          3),
        stdev = round(var ** 0.5,    3),
        min   = round(min(values),   3),
        max   = round(max(values),   3),
        count = n,
    )


def build_summary(records: list[dict]) -> list[dict]:
    """Aggregate per (label, case_id)."""
    from collections import defaultdict
    buckets: dict[tuple, list[dict]] = defaultdict(list)
    for r in records:
        if not r.get("error"):
            buckets[(r["label"], r["case_id"])].append(r)

    rows = []
    for (label, case_id), recs in sorted(buckets.items()):
        row = {"label": label, "case_id": case_id, "n_runs": len(recs)}
        for metric in SUMMARY_METRICS:
            vals = [r[metric] for r in recs if r.get(metric) is not None]
            s = _stats(vals)
            for stat, val in s.items():
                row[f"{metric}_{stat}"] = val
        rows.append(row)
    return rows

# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def write_excel(sources: list[tuple[Path, dict]], out_path: Path) -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    def _hfill(color):
        return PatternFill("solid", fgColor=color)

    def _font(color=None, bold=False, size=11):
        return Font(color=color or "000000", bold=bold, size=size,
                    name="Calibri")

    thin = Side(style="thin", color="D3D1C7")
    border = Border(bottom=thin)

    def style_header_row(ws, row_num: int, ncols: int) -> None:
        for col in range(1, ncols + 1):
            c = ws.cell(row=row_num, column=col)
            c.fill      = _hfill(HEADER_BG)
            c.font      = _font(HEADER_FG, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border    = border
        ws.row_dimensions[row_num].height = 30

    def autofit(ws) -> None:
        for col in ws.columns:
            width = max((len(str(c.value or "")) for c in col), default=8) + 4
            ws.column_dimensions[col[0].column_letter].width = min(width, 36)

    def freeze(ws, cell="B2") -> None:
        ws.freeze_panes = cell

    # ------------------------------------------------------------------
    # Sheet 1: metadata
    # ------------------------------------------------------------------
    ws_meta = wb.create_sheet("metadata")
    headers = METADATA_COLUMNS
    for col, h in enumerate(headers, 1):
        ws_meta.cell(row=1, column=col, value=h)
    style_header_row(ws_meta, 1, len(headers))

    for row_i, (src_path, data) in enumerate(sources, 2):
        records = data.get("records", [])
        errors  = sum(1 for r in records if r.get("error"))
        row = {
            "label":          data.get("label", src_path.stem),
            "schema_version": data.get("schema_version", "—"),
            "started_at":     data.get("started_at", "—"),
            "finished_at":    data.get("finished_at", "—"),
            "hostname":       data.get("hostname", "—"),
            "os":             data.get("os", "—"),
            "cpu_count":      data.get("cpu_count"),
            "python_version": data.get("python_version", "—"),
            "source_file":    str(src_path),
            "total_records":  len(records),
            "error_records":  errors,
        }
        for col, h in enumerate(headers, 1):
            ws_meta.cell(row=row_i, column=col, value=row.get(h))

    autofit(ws_meta)
    freeze(ws_meta, "A2")

    # ------------------------------------------------------------------
    # Sheet 2: raw
    # ------------------------------------------------------------------
    ws_raw = wb.create_sheet("raw")
    col_keys   = [k for k, _ in RAW_COLUMNS]
    col_labels = [l for _, l in RAW_COLUMNS]

    for col, label in enumerate(col_labels, 1):
        ws_raw.cell(row=1, column=col, value=label)
    style_header_row(ws_raw, 1, len(col_labels))

    all_records: list[dict] = []
    for src_path, data in sources:
        for rec in data.get("records", []):
            all_records.append(flatten_record(rec))

    for row_i, rec in enumerate(all_records, 2):
        is_error = bool(rec.get("error"))
        for col, key in enumerate(col_keys, 1):
            c = ws_raw.cell(row=row_i, column=col, value=rec.get(key))
            if is_error:
                c.fill = _hfill(ERROR_BG)
                c.font = _font(ERROR_FG)
            elif row_i % 2 == 0:
                c.fill = _hfill(ALT_ROW_BG)

    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(col_keys))}1"
    autofit(ws_raw)
    freeze(ws_raw, "C2")

    # ------------------------------------------------------------------
    # Sheet 3: summary
    # ------------------------------------------------------------------
    ws_sum = wb.create_sheet("summary")
    ok_records = [r for r in all_records if not r.get("error")]
    summary_rows = build_summary(ok_records)

    if summary_rows:
        sum_headers = list(summary_rows[0].keys())
        for col, h in enumerate(sum_headers, 1):
            ws_sum.cell(row=1, column=col, value=h)
        style_header_row(ws_sum, 1, len(sum_headers))

        for row_i, row in enumerate(summary_rows, 2):
            for col, key in enumerate(sum_headers, 1):
                c = ws_sum.cell(row=row_i, column=col, value=row.get(key))
                if row_i % 2 == 0:
                    c.fill = _hfill(ALT_ROW_BG)

        # colour scale on wall_ms_mean so slowest cases are obvious
        try:
            mean_col = sum_headers.index("wall_ms_mean") + 1
            col_letter = get_column_letter(mean_col)
            n = len(summary_rows)
            ws_sum.conditional_formatting.add(
                f"{col_letter}2:{col_letter}{n+1}",
                ColorScaleRule(
                    start_type="min",  start_color=GREEN_BG,
                    mid_type="percentile", mid_value=50, mid_color="FFFFFF",
                    end_type="max",    end_color="F8696B",
                ),
            )
        except ValueError:
            pass

        ws_sum.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}1"
        autofit(ws_sum)
        freeze(ws_sum, "C2")

    # ------------------------------------------------------------------
    # Finalize
    # ------------------------------------------------------------------
    wb.save(out_path)
    print(f"  wrote {out_path}")
    print(f"    sheets : metadata, raw, summary")
    print(f"    records: {len(all_records)} raw  |  {len(summary_rows)} summary rows")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("files", nargs="+", type=Path,
                   help="one or more results_*.json files")
    p.add_argument("--out", default="master.xlsx", type=Path,
                   help="output Excel file (default: master.xlsx)")
    args = p.parse_args()

    missing = [f for f in args.files if not f.exists()]
    if missing:
        print(f"files not found: {missing}", file=sys.stderr)
        sys.exit(1)

    sources = [(f, load_json(f)) for f in args.files]
    write_excel(sources, args.out)


if __name__ == "__main__":
    main()